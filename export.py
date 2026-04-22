"""
======================================================
Excel Export
======================================================
What this file does:
Takes the outputs from load_data.py, var,py, rolling_window.py, 
exceptions.py, and stress_test.py and writes them into a single
structurred Excel workbook.

Sheets produced: 
    1. README — methodology notes and column definitions
    2. SPX Returns — full cleaned return series 
    3. VaR Summary — rolling day-by-day VaR and exceptions 
    4. VaR Latest — single-row snapshot for the most recent date
    5. Stress Test — 9-scenario P&L grid 
    6. Return Stats — distributional statistics of the lookback window
    7. VaR Chart — rolling 20-day VaR time series 
    8. Distributions — return histograms 

Relies on:
    - load_data.py        
    - rolling_window.py   
    - var.py              
    - exceptions.py       
    - stress_test.py      
"""
import os
import sys
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")           
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from load_data import load_spx
from rolling_window import describe_window, default_lookback_years, \
        default_outlier_cutoff, default_MC_draws
from var import compute_var, default_confidence, default_horizon_days
from exceptions import run_backtest, compute_exception_stats
from stress_test import compute_stress_test
from read_positions import (find_drm_files, prompt_account_selection, 
                            load_account, prompt_scenario_parameters)

# ======================================================
# Configuration
# ======================================================
csv_filename = "spxtr_level_data.csv"
OUTPUT_FILE = "DRM_Analysis_Output.xlsx"
cache_file = "backtest_cache.csv"
chart_temp_dir = "_chart_temp" # temporary folder for chart images
worksheet_folder = "Worksheets"

# ======================================================
# Colour palette
# ======================================================
NAVY = "1F3864"
WHITE = "FFFFFF"
RED = "C00000"
GREEN = "375623"
L_RED = "FFCCCC"
L_GREEN = "E2EFDA"
L_BLUE = "DDEEFF"
GRAY = "F2F2F2"
D_GRAY = "D9D9D9"
ORANGE = "ED7D31"

# ======================================================
# Shared style helpers
# ======================================================
THIN_BORDER = Border(
    left = Side(style="thin", color="BFBFBF"),
    right = Side(style="thin", color="BFBFBF"),
    top = Side(style="thin", color="BFBFBF"),
    bottom = Side(style="thin", color="BFBFBF"),
)


def _hdr(ws, row: int, col: int, value, width: int = None):
    """Write a navy header cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row: int, col: int, value,
          fmt: str = None, bold: bool = False,
          color: str = "000000", bg: str = None,
          align: str = "center"):
    """Write a data cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(color=color, name="Arial", size=10, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


def _section_title(ws, row: int, col: int, value: str, n_cols: int = 1):
    """Write a section title spanning n_cols columns."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=NAVY, name="Arial", size=11)
    c.fill = PatternFill("solid", fgColor=L_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = THIN_BORDER
    if n_cols > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + n_cols - 1)
    ws.row_dimensions[row].height = 18
    return c


def _freeze(ws, cell: str = "A2"):
    ws.freeze_panes = cell


def _auto_width(ws, min_w: int = 10, max_w: int = 40):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = \
            min(max(length + 2, min_w), max_w)

# ======================================================
# Sheet 1 — README
# ======================================================
def _write_readme(wb: Workbook):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 75

    title = ws.cell(row = 1, column = 1, 
                    value = "DRM Analysis — Methodology and Column Guide")
    title.font = Font(bold = True, size = 14, name = "Arial", color = NAVY)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    content = [
        ("SECTION", "DESCRIPTION"),
        ("", ""),
        ("── INPUT DATA ──", ""),
        ("SPX Returns sheet", "Full SPX total-return index history. "
         "Returns = (Close_t / Close_{t-1}) - 1."),
        ("", ""),
        ("── VaR METHODOLOGY ──", ""),
        ("Lookback window", "Trailing 3 years (~755 trading days) "
         "of daily returns."),
        ("Confidence level", "99% — i.e. the 1st percentile of the "
         "return distribution."),
        ("HS-VaR", "1st percentile of the raw historical "
         "return distribution. No distributional "
         "assumptions."),
        ("MC-VaR", "1st percentile of returns resampled "
         "(with replacement) from the filtered "
         "pool. Outliers beyond ±3.5% excluded."),
        ("20-Day VaR", "1-Day VaR * √20  "
         "(square-root-of-time scaling)."),
        ("", ""),
        ("── VaR SUMMARY COLUMNS ──", ""),
        ("date", "Trading date."),
        ("actual_return", "Realized SPX daily return."),
        ("hs_var_1day", "HS 1-Day 99% VaR (negative = loss "
         "threshold)."),
        ("mc_var_1day", "MC 1-Day 99% VaR."),
        ("hs_var_scaled", "HS VaR scaled to the chosen horizon."),
        ("mc_var_scaled", "MC VaR scaled to the chosen horizon."),
        ("exception_hs", "1 if actual_return < hs_var_1day "
         "(breach), else 0."),
        ("exception_mc", "1 if actual_return < mc_var_1day, "
         "else 0."),
        ("", ""),
        ("—— STRESS TEST COLUMNS ──", ""),
        ("rm_code", "Scenario: RM01 = -40%, …, RM00=0%, "
         "…, RM08=+40%."),
        ("shock_pct", "Hypothetical index move applied to "
         "current SPX level."),
        ("spx_stressed", "SPX level after applying the shock."),
        ("pl_pct", "Portfolio P&L as % of NAV "
         "(= beta * shock)."),
        ("pl_dollars", "Dollar P&L = pl_pct * NAV."),
        ("position_value", "Portfolio value after shock "
         "= NAV + pl_dollars."),
        ("", ""),
        ("── EXCEPTION NOTES ──", ""),
        ("Expected rate", "At 99% confidence, ~1% of days should "
         "be exceptions."),
        ("Traffic light", "GREEN ≤ 1.5*, YELLOW ≤ 3*, RED > 3* "
         "the expected rate."),
        ("Backtesting", "Comparing observed exception rate to "
         "expected 1% validates model calibration."),
        ("Clustering", "Exceptions appearing in the same year "
         "signal regime failures, not random noise."),
    ]

    for i, (col_a, col_b) in enumerate(content, start = 2):
        a = ws.cell(row = i, column = 1, value = col_a)
        b = ws.cell(row = i, column = 2, value = col_b)
        if col_a == "SECTION":
            for c in (a, b):
                c.font = Font(bold = True, color = WHITE, name = "Arial", size = 10)
                c.fill = PatternFill("solid", fgColor = NAVY)
        elif col_a.startswith("——"):
            for c in (a, b):
                c.font = Font(bold = True, color = NAVY, name = "Arial", size = 10)
                c.fill = PatternFill("solid", fgColor = L_BLUE)
        else: 
            a.font = Font(bold = True, color = "000000", name = "Arial", size = 10)
            a.fill = PatternFill("solid", fgColor = GRAY)
            b.font = Font(color = "000000", name = "Arial", size = 10)
        b.alignment = Alignment(wrap_text = True, vertical = "top")
        ws.row_dimensions[i].height = 16

# ======================================================
# Sheet 2 — SPX Returns
# ======================================================
def _write_spx_returns(wb: Workbook, spx: pd.DataFrame): 
    ws = wb.create_sheet("SPX Returns")
    _section_title(ws, 1, 1, "SPX Total Return Index - Daily History", 3)

    headers = ["Date", "SPX Close", "Daily Return"]
    widths = [14, 14, 16]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, w)
    
    for r, row in enumerate(spx.itertuples(index = False), 3):
        bg = GRAY if r % 2 == 0 else None
        _cell(ws, r, 1, row.date.date(), fmt = "YYYY-MM-DD", bg = bg)
        _cell(ws, r, 2, row.close, fmt = "#,##0.00", bg = bg)
        ret_color = RED if row.daily_return < 0 else GREEN
        _cell(ws, r, 3, row.daily_return, fmt = "0.000%",
              color = ret_color, bg = bg)
    
    _freeze(ws, "A3")
    ws.auto_filter.ref = "A2:C2"

# ======================================================
# Sheet 3 — VaR Summary
# ======================================================
def _write_var_summary(wb: Workbook, bt: pd.DataFrame):
    ws = wb.create_sheet("VaR Summary")
    _section_title(ws, 1, 1, 
                   "Rolling VaR - Day-by-day Backtesting Results", 9)
    headers = ["Date", "SPX Close", "Actual Return", 
              "HS VaR 1D", "MC VaR 1D", "HS VaR Scaled", 
              "MC VaR Scaled", "Exception HS", "Exception MC"]
    widths = [13, 12, 15, 13, 13, 14, 14, 14, 14]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, w)

    for r, row in enumerate(bt.itertuples(index = False), 3):
        bg = GRAY if r % 2 == 0 else None
        _cell(ws, r, 1, row.date.date(), fmt = "YYY-MM-DD", bg = bg)
        _cell(ws, r, 2, row.spx_close if hasattr(row, "spx_close")
              else None, fmt = "#, ##0.00", bg = bg)
        _cell(ws, r, 3, row.actual_return, fmt = "0.000%", 
              color = RED if row.actual_return < 0 else GREEN, bg = bg)
        _cell(ws, r, 4, row.hs_var_1day, fmt="0.000%", color=RED, bg=bg)
        _cell(ws, r, 5, row.mc_var_1day, fmt="0.000%", color=RED, bg=bg)
        _cell(ws, r, 6, row.hs_var_scaled, fmt="0.000%", color=RED, bg=bg)
        _cell(ws, r, 7, row.mc_var_scaled, fmt="0.000%", color=RED, bg=bg)

        for col, exc in [(8, row.exception_hs), (9, row.exception_mc)]:
            if exc == 1:
                _cell(ws, r, col, "BREACH", color = RED, 
                      bold = True, bg = L_RED)
            else:
                _cell(ws, r, col, "OK", color = GREEN, bg = L_GREEN)
    
    _freeze(ws, "A3")
    ws.auto_filter.ref = "A2:I2"

# ======================================================
# Sheet 4 — VaR Latest
# ======================================================
def _write_var_latest(wb: Workbook, bt: pd.DataFrame, 
                      confidence: float, horizon_days: int, 
                      var_result: dict = None, 
                      account_meta: dict = None):
    
    ws = wb.create_sheet("VaR Latest")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    _section_title(ws, 1, 1, "VaR Snapshot - Most Recent Date", 3)

    latest = bt.iloc[-1]
    prev = bt.iloc[-2] if len(bt) > 1 else latest

    _hdr(ws, 2, 1, "Metric")
    _hdr(ws, 2, 2, f"T1 ({latest['date'].date()})")
    _hdr(ws, 2, 3, f"T0 ({prev['date'].date()})")

    pct_rows = {
        "HS-VaR 1-Day (99%)", "MC-VaR 1-Day (99%)", 
        f"HS-VaR {horizon_days}-Day", 
        f"MC-VaR {horizon_days}-Day", 
        "Actual Return"}
    
    rows_data = [
        ("Calculation Date",
         str(latest["date"].date()), str(prev["date"].date())),
        ("Actual Return",
         latest["actual_return"], prev["actual_return"]),
        ("HS-VaR 1-Day (99%)",
         latest["hs_var_1day"], prev["hs_var_1day"]),
        ("MC-VaR 1-Day (99%)",
         latest["mc_var_1day"], prev["mc_var_1day"]),
        (f"HS-VaR {horizon_days}-Day",
         latest["hs_var_scaled"], prev["hs_var_scaled"]),
        (f"MC-VaR {horizon_days}-Day",
         latest["mc_var_scaled"], prev["mc_var_scaled"]),
        ("Exception HS",
         "Breach" if latest["exception_hs"] else "Ok",
         "Breach" if prev["exception_hs"]   else "Ok"),
        ("Exception MC",
         "Breach" if latest["exception_mc"] else "Ok",
         "Breach" if prev["exception_mc"]   else "Ok"),
    ]

       # When fund VaR was computed, add rows below the SPX VaR rows
    if var_result and var_result.get("fund_var_1day") is not None:
        rows_data.append(("─" * 28, "", ""))
        rows_data.append(("Fund VaR 1-Day (ours)",
                          var_result["fund_var_1day"], None))
        rows_data.append((f"Fund VaR {horizon_days}-Day (ours)",
                          var_result["fund_var_scaled"], None))
        rows_data.append(("VaR Ratio (fund / SPX)",
                          var_result["fund_var_ratio"], None))
        if account_meta:
            rows_data.append(("─" * 28, "", ""))
            rows_data.append(("DRM Fund VaR 1-Day",
                              account_meta.get("drm_fund_var_1d"), None))
            rows_data.append((f"DRM Fund VaR 20-Day",
                              account_meta.get("drm_fund_var_20d"), None))
            rows_data.append(("DRM VaR Ratio",
                              account_meta.get("drm_var_ratio"), None))
        pct_rows.update({
            "Fund VaR 1-Day (ours)",
            f"Fund VaR {horizon_days}-Day (ours)",
            "DRM Fund VaR 1-Day", "DRM Fund VaR 20-Day",
        })

    for r, (label, t1, t0) in enumerate(rows_data, 3):
        bg = GRAY if r % 2 == 0 else None
        ws.cell(row = r, column = 1, value = label).font = \
        Font(bold = True, name = "Arial", size = 10)
        ws.cell(row = r, column = 1).fill = \
        PatternFill("solid", fgColor = D_GRAY)
        ws.cell(row = r, column = 1).border = THIN_BORDER

        fmt = "0.000%" if label in pct_rows else None
        for col, val in [(2, t1), (3, t0)]:
            if label.startswith("Exception"):
                exc_bg = L_RED if val == "BREACH" else L_GREEN
                exc_color = RED     if val == "BREACH" else GREEN
                _cell(ws, r, col, val, color=exc_color, bg = exc_bg)
            else:
                _cell(ws, r, col, val, fmt = fmt, bg = bg)

# ======================================================
# Sheet 5 — Stress test
# ======================================================
def _write_stress_test(wb: Workbook, stress_result: dict, 
                       positions: list = None):
    
    ws = wb.create_sheet("Stress test")
    ws.sheet_view.showGridLines = False

    spx_level = stress_result["spx_level"]
    nav = stress_result["portfolio_value"]
    beta = stress_result["beta"]
    date = stress_result["as_of_date"].date()

    _section_title(ws, 1, 1, 
                   f"Stress Test  | As of {date}  |  "
                   f"SPX: {spx_level:.2f}  |  "
                   f"NAV: ${nav:,.0f}  |  Beta: {beta}", 7)
    
    def write_grid(ws, start_row: int,
                   grid: pd.DataFrame, title: str):
        _section_title(ws, start_row, 1, title, 7)
        headers = ["Scenario", "Shock %", "SPX Current",
                   "SPX Stressed", "P&L %", "P&L ($)", "Position ($)"]
        widths  = [11, 10, 14, 14, 10, 15, 16]
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            _hdr(ws, start_row + 1, c, h, w)

        for r, row in enumerate(grid.itertuples(index=False),
                                start_row + 2):
            shock = row.shock_pct
            if shock < 0:
                bg, tc = L_RED, RED
            elif shock > 0:
                bg, tc = L_GREEN, GREEN
            else:
                bg, tc = L_BLUE,  NAVY
            _cell(ws, r, 1, row.rm_code, bold=True, color=NAVY, bg=bg)
            _cell(ws, r, 2, shock, fmt="0%", bg=bg, color=tc)
            _cell(ws, r, 3, row.spx_current, fmt="#,##0.00", bg=bg)
            _cell(ws, r, 4, row.spx_stressed, fmt="#,##0.00", bg=bg)
            _cell(ws, r, 5, row.pl_pct, fmt="+0.00%;-0.00%",
                  bg=bg, color=tc)
            _cell(ws, r, 6, row.pl_dollars,
                  fmt='#,##0;[Red]-#,##0;"-"', bg=bg, color=tc)
            _cell(ws, r, 7, row.position_value,  fmt="#,##0", bg=bg)

    # Grid 1 starts at row 3, Grid 2 starts 14 rows below Grid 1
    n_rows = len(stress_result["grid1"])
    write_grid(ws, 3, stress_result["grid1"], 
               "GRID 1 - Current Vols (market move only)")
    write_grid(ws, 3 + n_rows + 4, stress_result["grid2"],
               f"GRID 2 — Stressed Vols  "
               f"(market move + {stress_result['vol_multiplier']:.0f}* vol)")
    
    # Greeks sectn
    greeks_row = 3 + 2*(n_rows + 4) + 2
    _section_title(ws, greeks_row, 1, "Greeks - Base Case (RM00)", 3)
    base = stress_result["grid1"][stress_result["grid1"]["rm_code"] == "RM00"].iloc[0]

    greek_rows = [
        ("Delta ($ per SPX point)", base["delta"], "#,##0.00"),
        ("Gamma (delta Δ per SPX point)", base["gamma"], "#,##0.0000"),
        ("Vega ($ per 1 vol point)", base["vega"], "#,##0.00"),
        ("Theta ($ per trading day)", base["theta"], "#,##0.00"),
        ("Rho ($ per 1% rate move)", base["rho"], "#,##0.00"),
    ]

    for i, (label, val, fmt) in enumerate(greek_rows, greeks_row + 1):
        ws.cell(row = i, column = 1, value = label).font = \
            Font(bold = True, name = "Arial", size = 10)
        ws.cell(row = i, column = 1).fill = \
            PatternFill("solid", fgColor = GRAY)
        ws.cell(row = i, column = 1).border = THIN_BORDER
        ws.column_dimensions["A"].width = 34
        _cell(ws, i, 2, val, fmt = fmt)
    
    # Per-position Greeks — only when real positions are available
    if positions:
        from greeks import compute_greeks
        pos_greek_row = greeks_row + len(greek_rows) + 3
        _section_title(ws, pos_greek_row, 1,
                       "Position-Level Greeks (Base Case)", 7)
        pg_headers = ["Symbol", "Type", "Qty",
                      "Delta $", "Gamma $", "Vega $",
                      "Theta $", "Rho $"]
        pg_widths  = [22, 5, 10, 14, 14, 12, 12, 12]
        for c, (h, w) in enumerate(zip(pg_headers, pg_widths), 1):
            _hdr(ws, pos_greek_row + 1, c, h, w)

        for pr, p in enumerate(positions, pos_greek_row + 2):
            bg = GRAY if pr % 2 == 0 else None
            try:
                g = compute_greeks(
                    S=p["S"], K=p["K"], T_days=p["T_days"],
                    r=p["r"], sigma=p["sigma"],
                    option_type=p["option_type"],
                    quantity=p["quantity"], spc=p["spc"], q=p["q"])
                delta = g["delta_dollar"]
                gamma = g["gamma_dollar"]
                vega  = g["vega_dollar"]
                theta = g["theta_dollar"]
                rho   = g["rho_dollar"]
            except Exception:
                delta = gamma = vega = theta = rho = None
            qty_color = RED if p["quantity"] < 0 else GREEN
            _cell(ws, pr, 1, p["symbol"], bold=True, align="left", bg=bg)
            _cell(ws, pr, 2, p["option_type"].upper(), bg=bg)
            _cell(ws, pr, 3, p["quantity"], fmt="#,##0.00",
                  color=qty_color, bg=bg)
            _cell(ws, pr, 4, delta, fmt="#,##0", bg=bg)
            _cell(ws, pr, 5, gamma, fmt="#,##0.0000", bg=bg)
            _cell(ws, pr, 6, vega,  fmt="#,##0", bg=bg)
            _cell(ws, pr, 7, theta, fmt="#,##0.00", bg=bg)
            _cell(ws, pr, 8, rho,   fmt="#,##0.00", bg=bg)

# ======================================================
# Sheet 6 — Return Statistics
# ======================================================
def _write_return_stats(wb: Workbook, spx: pd.DataFrame, 
                        lookback_years: int, outlier_cutoff: float):
    
    ws = wb.create_sheet("Return Stats")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    latest_date  = spx["date"].max()
    cutoff = latest_date - pd.DateOffset(years=lookback_years)
    window_rets = spx[(spx["date"] > cutoff) &
                       (spx["date"] <= latest_date)]["daily_return"]
    full_rets = spx["daily_return"]

    _section_title(ws, 1, 1, "Return Distribution Statistics", 3)
    _hdr(ws, 2, 1, "Statistic")
    _hdr(ws, 2, 2, f"{lookback_years}-Year Window")
    _hdr(ws, 2, 3, "Full History")

    stats = [
        ("As-of Date", str(latest_date.date()), str(spx["date"].max().date())),
        ("Start Date", str(cutoff.date()), str(spx["date"].min().date())),
        ("Observations", len(window_rets), len(full_rets)),
        ("Mean Daily Return", window_rets.mean(), full_rets.mean()),
        ("Median Daily Return", window_rets.median(), full_rets.median()),
        ("Std Dev (Daily)", window_rets.std(), full_rets.std()),
        ("Min (Worst Day)", window_rets.min(), full_rets.min()),
        ("Max (Best Day)", window_rets.max(), full_rets.max()),
        ("1st Pctile (VaR)", window_rets.quantile(0.01), full_rets.quantile(0.01)),
        ("5th Percentile", window_rets.quantile(0.05), full_rets.quantile(0.05)),
        ("Kurtosis", window_rets.kurt(), full_rets.kurt()),
        ("Skewness", window_rets.skew(), full_rets.skew()),
        ("% Positive Days", (window_rets > 0).mean(), (full_rets > 0).mean()),
        ("% Negative Days", (window_rets < 0).mean(), (full_rets < 0).mean()),
        ("Annualised Return", window_rets.mean() * 252, full_rets.mean() * 252),
        ("Annualised Vol", window_rets.std() * 252**0.5, full_rets.std() * 252**0.5),
        (f"Outliers (|r|>{outlier_cutoff:.1%})",
         int((window_rets.abs() > outlier_cutoff).sum()),
         int((full_rets.abs() > outlier_cutoff).sum())),
    ]

    pct_labels = {"Mean Daily Return", "Median Daily Return", "Std Dev (Daily)",
                  "Min (Worst Day)", "Max (Best Day)", "1st Pctile (VaR)",
                  "5th Percentile", "% Positive Days", "% Negative Days",
                  "Annualised Return", "Annualised Vol"}
    
    for r, (label, win_val, full_val) in enumerate(stats, 3):
        bg = GRAY if r % 2 == 0 else None
        ws.cell(row = r, column = 1, value = label).font = \
            Font(bold = True, name = "Arial", size = 10)
        ws.cell(row = r, column = 1).fill = \
            PatternFill("solid", fgColor = D_GRAY if bg else GRAY)
        ws.cell(row = r, column = 1).border = THIN_BORDER
        fmt = "0.0000%" if label in pct_labels else \
              ("#,##0" if isinstance(win_val, (int, float)) and
               abs(win_val) > 1 else "0.0000")
        _cell(ws, r, 2, win_val,  fmt = fmt, bg = bg)
        _cell(ws, r, 3, full_val, fmt = fmt, bg = bg)

    
# ======================================================
# Chart helpers
# ======================================================
def _ensure_chart_dir():
    chart_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             chart_temp_dir)
    os.makedirs(chart_dir, exist_ok=True)
    return chart_dir

def _cleanup_chart_dir():
    import shutil
    chart_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             chart_temp_dir)
    if os.path.exists(chart_dir):
        shutil.rmtree(chart_dir)

def _save_fig(fig, filename: str) -> str:
    chart_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             chart_temp_dir)
    path = os.path.join(chart_dir, filename)
    fig.savefig(path, dpi = 150, bbox_inches = "tight",
                facecolor = "white", edgecolor = "none")
    plt.close(fig)
    return path

# ======================================================
# Sheet 7 — Rolling VaR Chart
# ======================================================
def _make_var_chart(bt: pd.DataFrame, 
                    horizon_days: int, 
                    confidence: float) -> str:
    """ 
    Replicate rollong VaR time series. 
    """
    fig, ax = plt.subplots(figsize = (13, 5))

    dates = bt["date"]
    var_pos = bt["hs_var_scaled"].abs()

    ax.plot(dates, var_pos, color = "#C00000", linewidth = 1.2, zorder = 3)

    # Shade crisis periods for context
    ax.set_title(
        f"SPX VaR  |  {confidence:.0%} confidence  |  "
        f"{horizon_days}-day horizon  |  3-year rolling window",
        fontsize=12, fontweight="bold", pad=12)
    ax.set_xlabel("Date", fontsize=10)
    ax.set_ylabel("VaR  (%)", fontsize=10)
    ax.yaxis.set_major_formatter(mticker.PercentFormatter(xmax=1, decimals=1))
    ax.set_xlim(dates.min(), dates.max())
    ax.set_ylim(0, var_pos.max() * 1.10)
    ax.grid(True, linestyle="--", alpha=0.4, zorder=0)
    ax.spines[["top", "right"]].set_visible(False)

    # Annotate the peak (financial crisis)
    peak_idx = var_pos.idxmax()
    ax.annotate(
        f"Peak: {var_pos[peak_idx]:.1%}\n{bt['date'][peak_idx].strftime('%b %Y')}",
        xy=(bt["date"][peak_idx], var_pos[peak_idx]),
        xytext=(30, -25), textcoords="offset points",
        arrowprops=dict(arrowstyle="->", color="black", lw=0.8),
        fontsize=8, color="black")
    
    fig.tight_layout()
    return _save_fig(fig, "var_timeseries.png")

def _write_var_chart(wb: Workbook, chart_path: str):
    ws = wb.create_sheet("VaR Chart")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, 1,
                   "Rolling VaR Time Series — SPX  "
                   "(3-year window, 99% confidence)", 1)
    img = XLImage(chart_path)
    img.anchor = "A3"
    ws.add_image(img)

# ======================================================
# Sheet 8 — Return Distributions
# ======================================================
def _make_distribution_charts(spx: pd.DataFrame, 
                              window_data: dict, 
                              mc_simulated: np.ndarray) -> list:
    """
    Produce the histogram panels. 
    """
    paths = []

    full_rets = spx["daily_return"].values
    window_rets = window_data["full_window"].values
    pool_rets = window_data["filtered_pool"].values
    n_window = len(window_rets)
    n_full = len(full_rets)
    n_mc = len(mc_simulated)
    cutoff = window_data["outlier_cutoff"]

    # --- 3-year window ---
    fig, ax = plt.subplots(figsize=(6.5, 4))
    ax.hist(window_rets, bins=60, color="#1F77B4", edgecolor="white",
            linewidth=0.3, alpha=0.85)
    var_1pct = float(np.percentile(window_rets, 1))
    ax.axvline(var_1pct, color="#C00000", linewidth=1.5,
               linestyle="--", label=f"1st pctile: {var_1pct:.3f}")
    ax.set_title(
        f"SPX {window_data['lookback_years']}Y Window — "
        f"{n_window:,} Daily Returns\n"
        f"[{window_rets.min():.4f},  {window_rets.max():.4f}]  "
        f"μ={window_rets.mean():.5f}  σ={window_rets.std():.5f}",
        fontsize=9)
    ax.set_xlabel("Daily Return", fontsize=9)
    ax.set_ylabel("Frequency", fontsize=9)
    ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1, decimals=1))
    ax.legend(fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.3)
    fig.tight_layout()
    paths.append(_save_fig(fig, "dist_window.png"))

    # --- MC simulated distribution ---
    fig, ax = plt.subplots(figsize=(6.5, 4))
    ax.hist(mc_simulated, bins=60, color="#1F77B4", edgecolor="white",
            linewidth=0.3, alpha=0.85)
    mc_var = float(np.percentile(mc_simulated, 1))
    ax.axvline(mc_var, color="#C00000", linewidth=1.5,
               linestyle="--", label=f"MC VaR (1%): {mc_var:.3f}")

    # Show left tail values above the chart 
    left_tail = np.sort(mc_simulated)[:10]
    tail_str  = ",  ".join(f"{v:.5f}" for v in left_tail)
    ax.set_title(
        f"MC Distribution — {n_mc:,} Simulated Returns\n"
        f"[{mc_simulated.min():.4f},  {mc_simulated.max():.4f}]  "
        f"μ={mc_simulated.mean():.5f}  σ={mc_simulated.std():.5f}\n"
        f"Left tail: {tail_str[:60]}...",
        fontsize=8)
    ax.set_xlabel("Simulated Return", fontsize=9)
    ax.set_ylabel("Frequency", fontsize=9)
    ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1, decimals=1))
    ax.legend(fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.3)
    fig.tight_layout()
    paths.append(_save_fig(fig, "dist_mc.png"))

    # --- Full SPX history ---
    fig, ax = plt.subplots(figsize=(6.5, 4))
    ax.hist(full_rets, bins=100, color="#1F77B4", edgecolor="white",
            linewidth=0.2, alpha=0.85)
    full_var = float(np.percentile(full_rets, 1))
    ax.axvline(full_var, color="#C00000", linewidth=1.5,
               linestyle="--", label=f"1st pctile: {full_var:.3f}")
    n_out_d = int((full_rets < -cutoff).sum())
    n_out_u = int((full_rets >  cutoff).sum())
    ax.set_title(
        f"SPX Full History — {n_full:,} Daily Returns\n"
        f"[{full_rets.min():.4f},  {full_rets.max():.4f}]  "
        f"μ={full_rets.mean():.5f}  σ={full_rets.std():.5f}\n"
        f"Outliers (|r|>{cutoff:.1%}): down={n_out_d}  up={n_out_u}",
        fontsize=9)
    ax.set_xlabel("Daily Return", fontsize=9)
    ax.set_ylabel("Frequency", fontsize=9)
    ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1, decimals=1))
    ax.legend(fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.3)
    fig.tight_layout()
    paths.append(_save_fig(fig, "dist_full.png"))

    # --- Filtered pool/MCHist ---
    fig, ax = plt.subplots(figsize=(6.5, 4))
    ax.hist(pool_rets, bins=60, color="#1F77B4", edgecolor="white",
            linewidth=0.3, alpha=0.85)
    pool_var = float(np.percentile(pool_rets, 1))
    ax.axvline(pool_var, color="#C00000", linewidth=1.5,
               linestyle="--", label=f"1st pctile: {pool_var:.3f}")
    n_out_pool = int((np.abs(window_rets) > cutoff).sum())
    ax.set_title(
        f"SPX Filtered Pool — {len(pool_rets):,} Returns\n"
        f"(3Y window, outliers |r|>{cutoff:.1%} removed — "
        f"{n_out_pool} excluded)\n"
        f"[{pool_rets.min():.4f},  {pool_rets.max():.4f}]  "
        f"μ={pool_rets.mean():.5f}  σ={pool_rets.std():.5f}",
        fontsize=9)
    ax.set_xlabel("Daily Return", fontsize=9)
    ax.set_ylabel("Frequency", fontsize=9)
    ax.xaxis.set_major_formatter(mticker.PercentFormatter(xmax=1, decimals=1))
    ax.legend(fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.3)
    fig.tight_layout()
    paths.append(_save_fig(fig, "dist_pool.png"))

    return paths

def _write_distributions(wb: Workbook, chart_paths: list):
    ws = wb.create_sheet("Distributions")
    ws.sheet_view.showGridLines = False

    titles = [
        (1, "A3", "3-Year Window — Historical Returns"),
        (1, "K3", "MC Simulated Distribution"),
        (30, "A3", "Full SPX History (all years)"),
        (30, "K3", "Filtered Pool (outliers removed)"),
    ]

    anchors = ["A3", "K3", "A33", "K33"]
    labels  = [
        "3-Year Window — Historical Returns",
        "MC Simulated Distribution",
        "Full SPX History (all years)",
        "Filtered Pool (outliers removed)",
    ]

    _section_title(ws, 1, 1,
                   "Return Distributions — Historical and Simulated", 20)

    for path, anchor, label in zip(chart_paths, anchors, labels):
        img = XLImage(path)
        img.anchor = anchor
        ws.add_image(img)

# ======================================================
# Sheet 9 — Positions  (account mode only)
# ======================================================
def _write_positions(wb: Workbook, positions: list, account: str):
    """
    Write the 19 base positions with the computed Greeks
    side by side with DRM workbook's Greeks.
    """
    from greeks import compute_greeks

    ws = wb.create_sheet("Positions")
    ws.sheet_view.showGridLines = False

    _section_title(ws, 1, 1, 
                   f"Option Positions — {account.upper()}  "
                   f"(VAR0000 Base Case, MAGS excluded)", 14)
    
    headers = ["Symbol", "Type", "Qty", "Underlying $",
               "Strike", "T Days", "Sigma",
               "Current THV", "DRM THV", "THV Diff",
               "Current Delta $", "DRM Delta $",
               "Current Vega", "DRM Vega",
               "Current Theta", "DRM Theta",
               "Current Rho", "DRM Rho"]
    
    widths = [22, 5, 10, 13, 10, 8, 7,
              12, 12, 10, 13, 13,
              11, 11, 11, 11, 11, 11]
    
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, w)
    
    for r, p in enumerate(positions, 3):
        bg = GRAY if r% 2 == 0 else None

        # Compute Greeks
        try:
            g = compute_greeks(
                S = p["S"], 
                K = p["K"], 
                T_days = p["T_days"], 
                r = p["r"], 
                sigma = p["sigma"], 
                option_type = p["option_type"], 
                quantity = p["quantity"], 
                spc = p["spc"], 
                q = p["q"], 
            )
            our_thv   = g["bs_price"]
            our_delta = g["delta_dollar"]
            our_vega  = g["vega_dollar"]
            our_theta = g["theta_dollar"]
            our_rho   = g["rho_dollar"]
        except Exception:
            our_thv = our_delta = our_vega = our_theta = our_rho =  None
        
        drm_thv   = p.get("drm_thv")
        drm_delta = p.get("drm_delta_dollars")
        drm_vega  = p.get("drm_vega")
        drm_theta = p.get("drm_theta")
        drm_rho   = p.get("drm_rho")

        thv_diff = (our_thv - drm_thv
                    if our_thv is not None and drm_thv is not None
                    else None)
        
        qty_color = RED if p["quantity"] < 0 else GREEN

        _cell(ws, r,  1, p["symbol"], bold=True, align="left", bg=bg)
        _cell(ws, r,  2, p["option_type"].upper(), bg=bg)
        _cell(ws, r,  3, p["quantity"], fmt="#,##0.00",
              color=qty_color, bg=bg)
        _cell(ws, r,  4, p["S"], fmt="#,##0.00", bg=bg)
        _cell(ws, r,  5, p["K"], fmt="#,##0.00", bg=bg)
        _cell(ws, r,  6, p["T_days"], fmt="0.00", bg=bg)
        _cell(ws, r,  7, p["sigma"], fmt="0.000", bg=bg)
        _cell(ws, r,  8, our_thv, fmt="#,##0.0000", bg=bg)
        _cell(ws, r,  9, drm_thv, fmt="#,##0.0000", bg=bg)
        _cell(ws, r, 10, thv_diff, fmt="+#,##0.0000;-#,##0.0000",
              color=RED if thv_diff and abs(thv_diff) > 1 else "000000",
                                                                      bg=bg)
        _cell(ws, r, 11, our_delta, fmt="#,##0", bg=bg)
        _cell(ws, r, 12, drm_delta, fmt="#,##0", bg=bg)
        _cell(ws, r, 13, our_vega, fmt="#,##0", bg=bg)
        _cell(ws, r, 14, drm_vega, fmt="#,##0", bg=bg)
        _cell(ws, r, 15, our_theta, fmt="#,##0.00", bg=bg)
        _cell(ws, r, 16, drm_theta, fmt="#,##0.00", bg=bg)
        _cell(ws, r, 17, our_rho,   fmt="#,##0.00", bg=bg)
        _cell(ws, r, 18, drm_rho,   fmt="#,##0.00", bg=bg)
    
     # Note row
    note_row = len(positions) + 4
    note = ws.cell(row=note_row, column=1,
                   value="Note: MAGS positions excluded (non-SPX underlying). "
                         "The current THV uses Black-Scholes with q=0 (no dividend yield).")
    note.font = Font(italic=True, color="666666", name="Arial", size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row = note_row, start_column = 1,
                   end_row = note_row, end_column = 18)

# ======================================================
# Sheet 10 — Fund VaR (account mode only)
# ======================================================
def _make_fund_var_chart(var_result: dict, 
                         account: str, 
                         horizon_days: int) -> str:
    """
    Build the fund return distribution histogram with VaR marked.
    Returns the path to the saved PNG
    """
    dist = var_result["fund_var_dist"]
    fv1 = var_result["fund_var_1day"]

    fig, ax = plt.subplots(figsize = (10, 4.5))

    ax.hist(dist.values, bins = 60, color = "#1F77B4", 
            edgecolor = "white", linewidth = 0.3, alpha = 0.85)
    ax.axvline(fv1, color = "#C00000", linewidth = 1.8, 
               linestyle = "--", 
               label = f"Fund VaR (1%): {fv1:.4%}")
    ax.set_title(
        f"{account.upper()} — Fund Return Distribution  "
        f"({len(dist):,} scenarios)\n"
        f"Range: [{dist.min():.4f},  {dist.max():.4f}]  "
        f"Mean={dist.mean():.5f}  Std={dist.std():.5f}",
        fontsize=10, fontweight="bold")
    ax.set_xlabel("Fund Return", fontsize=9)
    ax.set_ylabel("Frequency",   fontsize=9)
    ax.xaxis.set_major_formatter(
        mticker.PercentFormatter(xmax=1, decimals=2))
    ax.legend(fontsize=9)
    ax.spines[["top", "right"]].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.35)
    fig.tight_layout()

    return _save_fig(fig, "fund_var_dist.png")

def _write_fund_var(wb: Workbook, var_result: dict, 
                    account_meta: dict, horizon_days: int):
    """
    Write the Fund VaR sheet: comparison table + distribution chart
    """
    ws = wb.create_sheet("Fund VaR")
    ws.sheet_view.showGridlines = False

    account = account_meta.get("account", "").upper()
    _section_title(ws, 1,1, 
                   f"Fund-Level VaR — {account} "
                   f"(Black-Scholes Repricing, Full History Scenarios)", 6)
    
    # --- Comparoson table ---
    _section_title(ws, 3, 1, "VaR Comparison 0ù Current model vs DRM Workbook", 6)

    headers = ["Metric", "Current Model", "DRM Workbook", "Difference", 
                "Direction", "Note"]
    widths = [30, 18, 18, 14, 14, 30]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 4, c, h, w)
    
    fv1 = var_result["fund_var_1day"]
    fvs = var_result["fund_var_scaled"]
    rat = var_result["fund_var_ratio"]
    hs1 = var_result["hs_var_1day"]
    hss = var_result["hs_var_scaled"]

    drm_fv1 = account_meta.get("drm_fund_var_1d")
    drm_fvs = account_meta.get("drm_fund_var_20d")
    drm_rat = account_meta.get("drm_var_ratio")
    drm_iv1 = account_meta.get("drm_idx_var_1d")
    drm_ivs = account_meta.get("drm_idx_var_20d")

    def diff_color(d):
        return RED if d and d < 0 else GREEN
    
    rows = [
       ("Fund VaR 1-Day",
         fv1, drm_fv1,
         fv1 - drm_fv1 if drm_fv1 else None,
         "0.000%"),
        (f"Fund VaR {horizon_days}-Day",
         fvs, drm_fvs,
         fvs - drm_fvs if drm_fvs else None,
         "0.000%"),
        ("SPX Index VaR 1-Day",
         hs1, drm_iv1,
         hs1 - drm_iv1 if drm_iv1 else None,
         "0.000%"),
        (f"SPX Index VaR {horizon_days}-Day",
         hss, drm_ivs,
         hss - drm_ivs if drm_ivs else None,
         "0.000%"),
        ("VaR Ratio (Fund / SPX)",
         rat, drm_rat,
         rat - drm_rat if drm_rat else None,
         "0.0000"), 
    ]

    for r, (label, ours, drm, diff, fmt) in enumerate(rows, 5):
        bg = GRAY if r % 2 == 0 else None
        _cell(ws, r, 1, label, bold=True, align="left", bg=bg)
        _cell(ws, r, 2, ours, fmt=fmt, color=RED, bg=bg)
        _cell(ws, r, 3, drm, fmt=fmt, color=RED, bg=bg)
        _cell(ws, r, 4, diff, fmt="+0.000%;-0.000%",
              color=diff_color(diff), bg=bg)
        
        # Direction 
        if ours is not None and drm is not None:
            direction = "Current < DRM" if abs(ours) < abs(drm) else "Current > DRM"
        else:
            direction = "—"
        _cell(ws, r, 5, direction, bg = bg)

        # --- Distribution Stats ---
        dist = var_result["fund_var_dist"]
        stat_row = note_row = 7 //2 + 3
        _section_title(ws, stat_row, 1, 
                       "Fund Return Distribution Statistics", 4)
        stats = [
            ("Scenarios", len(dist), "#,##0"),
            ("Positions repriced", var_result["fund_n_scenarios"], "#,##0"),
            ("Portfolio value", var_result["fund_portfolio_value"], "$#,##0"),
            ("SPY multiplier", var_result["fund_spy_multiplier"], "0.0000"),
            ("Min fund return", dist.min(), "0.0000%"),
            ("Max fund return", dist.max(), "0.0000%"),
            ("Mean fund return", dist.mean(), "0.0000%"),
            ("Std dev", dist.std(), "0.0000%"),
            ("1st percentile", float(np.percentile(dist, 1)), "0.0000%"),
        ]

        for i, (label, val, fmt) in enumerate(stats, stat_row + 1):
            bg = GRAY if i % 2 == 0 else None
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, name="Arial", size=10)
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=D_GRAY if bg else GRAY)
            ws.cell(row=i, column=1).border = THIN_BORDER
            _cell(ws, i, 2, val, fmt = fmt, bg=bg)
        
        return ws      

# ======================================================
# Sheet 11 — VaR Detail  (account mode only)
# ======================================================
def _write_var_detail(wb: Workbook, positions: list,
                      scenario_levels: "pd.Series",
                      base_spx: float, account: str):
    """
    For each scenario * position: show the repriced option value
    and P&L. Replicates the DRM VaR Detail sheet structure.
    """
    from greeks import bs_price as _bs

    ws = wb.create_sheet("VaR Detail")
    ws.sheet_view.showGridLines = False

    spy_prices = [p["S"] for p in positions if p["underlying"] == "spy"]
    spy_mult = base_spx / spy_prices[0] if spy_prices else 10.029
    net_posn = sum(p["drm_posn_dollars"] for p in positions
                        if p["drm_posn_dollars"] is not None)
    portfolio_val = abs(net_posn) if net_posn else 1.0

    _section_title(ws, 1, 1,
                   f"VaR Detail — {account.upper()}  "
                   f"(Base SPX: {base_spx:,.2f}  |  "
                   f"Portfolio: ${portfolio_val:,.0f})", 9)

    headers = ["Scenario", "SPX Level", "SPX Return",
               "Symbol", "Type", "Base Price",
               "New Price", "PL ($)", "PL (%)"]
    widths = [12, 12, 11, 24, 5, 12, 12, 14, 10]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, w)

    row = 3
    for i, spx_level in enumerate(scenario_levels):
        scenario_id = f"VAR{i+1:04d}VR"
        spx_ret = float(spx_level) / base_spx - 1.0
        scenario_pl = 0.0

        for p in positions:
            S_new = (float(spx_level) if p["underlying"] in ("spx","xsp")
                     else float(spx_level) / spy_mult)
            T_new = p["T_days"] - 1.0
            if T_new <= 0:
                new_price = (max(S_new - p["K"], 0.0)
                             if p["option_type"] == "call"
                             else max(p["K"] - S_new, 0.0))
            else:
                try:
                    new_price = _bs(S=S_new, K=p["K"],
                                    T=T_new/252.0, r=p["r"],
                                    sigma=p["sigma"],
                                    option_type=p["option_type"],
                                    q=p["q"])
                except Exception:
                    new_price = p.get("base_price", p["S"])

            base_price = p.get("base_price", p["S"])
            pl_pos = (new_price - base_price) * p["quantity"] * p["spc"]
            scenario_pl += pl_pos
            pl_pct = pl_pos / portfolio_val

            bg = None
            tc = RED   if pl_pos < 0 else GREEN

            _cell(ws, row, 1, scenario_id, bg=bg)
            _cell(ws, row, 2, float(spx_level), fmt="#,##0.00", bg=bg)
            _cell(ws, row, 3, spx_ret, fmt="+0.000%;-0.000%",
                  color=(RED if spx_ret < 0 else GREEN), bg=bg)
            _cell(ws, row, 4, p["symbol"], bold=True, align="left", bg=bg)
            _cell(ws, row, 5, p["option_type"].upper(), bg=bg)
            _cell(ws, row, 6, base_price, fmt="#,##0.0000", bg=bg)
            _cell(ws, row, 7, new_price, fmt="#,##0.0000", bg=bg)
            _cell(ws, row, 8, pl_pos, fmt='#,##0;[Red]-#,##0',
                  color=tc, bg=bg)
            _cell(ws, row, 9, pl_pct, fmt="+0.000%;-0.000%",
                  color=tc, bg=bg)
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:I2"  

 # ======================================================
# Sheet 12 — Stress Test Full Detail  (account mode only)
# ======================================================
def _write_stress_full_detail(wb: Workbook, positions: list,
                               stress_result: dict, account: str):
    """
    For each stress scenario * position: repriced value and P&L.
    Replicates the DRM Stress Test Full Detail sheet.
    """
    from greeks import bs_price as _bs, compute_greeks

    ws = wb.create_sheet("Stress Full Detail")
    ws.sheet_view.showGridLines = False

    base_spx = stress_result["spx_level"]
    spy_prices = [p["S"] for p in positions if p["underlying"] == "spy"]
    spy_mult = base_spx / spy_prices[0] if spy_prices else 10.029
    net_posn = sum(p["drm_posn_dollars"] for p in positions
                     if p["drm_posn_dollars"] is not None)
    portfolio_val = abs(net_posn) if net_posn else 1.0

    _section_title(ws, 1, 1,
                   f"Stress Test Full Detail — {account.upper()}  "
                   f"(Base SPX: {base_spx:,.2f}  |  "
                   f"Portfolio: ${portfolio_val:,.0f})", 11)

    headers = ["Scenario", "Shock %", "SPX Current", "SPX Stressed",
               "Symbol", "Type", "Base Price", "New Price",
               "PL ($)", "PL (%)", "Delta $"]
    widths = [10, 8, 12, 12, 24, 5, 12, 12, 14, 10, 13]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 2, c, h, w)

    row = 3
    for scenario in stress_result["scenarios"]:
        rm_code = scenario["rm_code"]
        shock = scenario["shock"]
        spx_stressed = base_spx * (1 + shock)

        if shock < 0:
            bg, tc = None, RED
        elif shock > 0:
            bg, tc = None, GREEN
        else:
            bg, tc = None, NAVY

        for p in positions:
            S_new = (spx_stressed if p["underlying"] in ("spx","xsp")
                     else spx_stressed / spy_mult)
            T_new = p["T_days"] - 1.0
            if T_new <= 0:
                new_price = (max(S_new - p["K"], 0.0)
                             if p["option_type"] == "call"
                             else max(p["K"] - S_new, 0.0))
            else:
                try:
                    new_price = _bs(S=S_new, K=p["K"],
                                    T=T_new/252.0, r=p["r"],
                                    sigma=p["sigma"],
                                    option_type=p["option_type"],
                                    q=p["q"])
                except Exception:
                    new_price = p.get("base_price", p["S"])

            base_price = p.get("base_price", p["S"])
            pl_pos = (new_price - base_price) * p["quantity"] * p["spc"]
            pl_pct = pl_pos / portfolio_val
            pl_color = RED if pl_pos < 0 else GREEN

            # Delta at the stressed price
            try:
                g = compute_greeks(S=S_new, K=p["K"], T_days=T_new,
                                   r=p["r"], sigma=p["sigma"],
                                   option_type=p["option_type"],
                                   quantity=p["quantity"],
                                   spc=p["spc"], q=p["q"])
                delta_stressed = g["delta_dollar"]
            except Exception:
                delta_stressed = None

            _cell(ws, row, 1, rm_code, bold=True, color=NAVY, bg=bg)
            _cell(ws, row, 2, shock, fmt="+0%;-0%",
                  color=tc, bg=bg)
            _cell(ws, row, 3, base_spx, fmt="#,##0.00", bg=bg)
            _cell(ws, row, 4, spx_stressed, fmt="#,##0.00", bg=bg)
            _cell(ws, row, 5, p["symbol"], bold=True,
                  align="left", bg=bg)
            _cell(ws, row, 6, p["option_type"].upper(), bg=bg)
            _cell(ws, row, 7, base_price, fmt="#,##0.0000", bg=bg)
            _cell(ws, row, 8, new_price, fmt="#,##0.0000", bg=bg)
            _cell(ws, row, 9, pl_pos,
                  fmt='#,##0;[Red]-#,##0', color=pl_color, bg=bg)
            _cell(ws, row, 10, pl_pct,
                  fmt="+0.000%;-0.000%", color=pl_color, bg=bg)
            _cell(ws, row, 11, delta_stressed, fmt="#,##0", bg=bg)
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:K2"
        
# ======================================================
# Cache helpers 
# ======================================================
def load_or_run_backtest(spx: pd.DataFrame, 
                         lookback_years: int, 
                         outlier_cutoff: float, 
                         mc_draws: int, 
                         horizon_days: int, 
                         confidence: float, 
                         cache_path: str) -> pd.DataFrame:
    """
    Load the backtest DataFrame from cache if it exists, 
    otherwise run the full rolling backtest and save to cache. 

    The cache filename encodes the key parameters so that changing 
    any parameter automatically triggers a fresh recalculation.
    """
    # encode parameters into cache filename
    param_tag = (f"ly{lookback_years}_oc{int(outlier_cutoff*1000)}"
                 f"_mc{mc_draws}_h{horizon_days}"
                 f"_c{int(confidence*100)}")
    base, ext = os.path.splitext(cache_path)
    tagged_cache = f"{base}_{param_tag}{ext}"

    if os.path.exists(tagged_cache):
        print(f"[Cache] Loading backtest from {tagged_cache}")
        bt = pd.read_csv(tagged_cache, parse_dates = ["date"])
        print(f"[Cache] Loaded {len(bt):,} rows.")
        return bt
    
    print(f"[Cache] No cache found. Running full backtest...")
    print(f"This takes ~2 minutes. Results will be cached for next run.")
    bt = run_backtest(
        spx,
        lookback_years = lookback_years,
        outlier_cutoff = outlier_cutoff,
        mc_draws = mc_draws,
        horizon_days = horizon_days,
        confidence = confidence,
    )

    # Add spx_close to the backtest DataFrame for sheet 3
    bt = bt.merge(
        spx[["date", "close"]].rename(columns = {"close": "spx_close"}), 
        on = "date", how = "left")
    
    bt.to_csv(tagged_cache, index = False)
    print(f"[Cache] Saved to {tagged_cache}")
    return bt

# ======================================================
# Export function 
# ======================================================
def export(spx: pd.DataFrame, 
           lookback_years: int = default_lookback_years, 
           outlier_cutoff: float = default_outlier_cutoff, 
           mc_draws: int = default_MC_draws, 
           horizon_days: int = default_horizon_days, 
           confidence: float = default_confidence, 
           portfolio_value: float = 10_000_000.0, 
           beta: float = 1.0, 
           vol_multiplier: float = 2.0, 
           output_path: str = None, 
           cache_path: str = None, 
           account_data: dict = None):
    """
    Run the full export pipeline and write the Excel workbook.

    Parameters: 
    - spx: full SPX DataFrame from load_spx()
    - lookback_years: VaR lookback window in years
    - outlier_cutoff: MC outlier filter threshold
    - mc_draws: number of MC draws
    - horizon_days: VaR scaling horizon in trading days
    - confidence: VaR confidence level
    - portfolio_value: portfolio NAV for stress test
    - beta: portfolio beta for stress test
    - vol_multiplier: vol scaling for stress test Grid 2
    - output_path: path for the output .xlsx file
    - cache_path: path for the backtest cache .csv file
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = output_path or os.path.join(script_dir, OUTPUT_FILE)
    cache_path = cache_path or os.path.join(script_dir, cache_file)
    
    _ensure_chart_dir() 

    # --- window data for the most recent date ---
    print("\n[Export] Building window data...")
    latest_date = spx["date"].max()
    window_data = describe_window(
        spx, latest_date, lookback_years, outlier_cutoff, mc_draws)

    # --- VaR for the most recent date ---
    print("[Export] Computing latest VaR...")
    positions = account_data["positions"] if account_data else None
    scenario_levels = account_data["scenario_levels"] if account_data else None
    base_spx = account_data["base_spx"] if account_data else None

    var_result = compute_var(
        window_data, 
        confidence = confidence, 
        horizon_days = horizon_days, 
        positions = positions, 
        scenario_levels = scenario_levels, 
        base_spx = base_spx, 
    )
    mc_simulated = var_result["mc_simulated"]

    # --- Rolling backtest (cached) ---
    bt = load_or_run_backtest(
        spx, lookback_years, outlier_cutoff,
        mc_draws, horizon_days, confidence, cache_path)

    # --- Stress test ---
    print("[Export] Running stress test...")
    stress_result = compute_stress_test(
        spx,
        portfolio_value = portfolio_value,
        beta = beta,
        vol_multiplier = vol_multiplier,
    )

    # --- Build charts ---
    print("[Export] Generating charts...")
    var_chart_path  = _make_var_chart(bt, horizon_days, confidence)
    dist_chart_paths = _make_distribution_charts(
        spx, window_data, mc_simulated)

    # --- Build workbook ---
    print("[Export] Building workbook...")
    wb = Workbook()
    wb.remove(wb.active)        # remove default blank sheet

    _write_readme(wb)
    _write_spx_returns(wb, spx)
    _write_var_summary(wb, bt)
    _write_var_latest(wb, bt, confidence, horizon_days, 
                      var_result = var_result, 
                      account_meta = account_data["metadata"] if account_data else None)
    _write_stress_test(wb, stress_result, 
                       positions = account_data["positions"] if account_data else None)
    _write_return_stats(wb, spx, lookback_years, outlier_cutoff)
    _write_var_chart(wb, var_chart_path)
    _write_distributions(wb, dist_chart_paths)

     # Account-specific sheets — only when an account was selected
    if account_data and var_result.get("fund_var_1day") is not None:
        account_name = account_data["metadata"]["account"]
        print("[Export] Writing Positions sheet...")
        _write_positions(wb, 
                         account_data["positions"], 
                         account_name)

        print("[Export] Writing Fund VaR sheet...")
        fund_chart_path = _make_fund_var_chart(
            var_result, account_name, horizon_days)
        ws_fv = _write_fund_var(
            wb, var_result, account_data["metadata"], horizon_days)
        img = XLImage(fund_chart_path)
        img.anchor = "A30"
        ws_fv.add_image(img)

        print("[Export] Writing VaR Detail sheet...")
        _write_var_detail(wb,
                          account_data["positions"],
                          account_data["scenario_levels"],
                          account_data["base_spx"],
                          account_name)

        print("[Export] Writing Stress Full Detail sheet...")
        _write_stress_full_detail(wb,
                                  account_data["positions"],
                                  stress_result,
                                  account_name)

    wb.save(output_file)
    print(f"[Export] Workbook saved -> {output_file}")

    _cleanup_chart_dir()
    print("[Export] Temporary chart files cleaned up.")
    return output_file

# ======================================================
# Entry Point
# ======================================================
if __name__ == "__main__":

    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(script_dir, csv_filename)
    worksheet_dir = os.path.join(script_dir, worksheet_folder)

    print("\n" + "=" * 65)
    print("DRM ANALYSIS — EXCEL EXPORT")
    print("=" * 65)

    spx = load_spx(csv_path)
    print(f"SPX data loaded: {len(spx):,} days "
          f"({spx['date'].min().date()} -> {spx['date'].max().date()})")
    
    # --- Account Selection ---
    selected = prompt_account_selection(worksheet_dir)

    account_data = None
    if selected:
        # Ask for scenario generation parameters
        n_secenarios, outlier_cutoff = prompt_scenario_parameters()
        account_data = load_account(
            filepath = selected["filepath"], 
            spx = spx, 
            n_scenarios = n_secenarios, 
            outlier_cutoff = outlier_cutoff, 
            seed = 123
        )
    else:
        print("\n Running in SPX-only mode.")
    
    # Run all export pipeline
    output_path = export(spx, account_data = account_data)

    print("\n" + "=" * 65)
    print("EXPORT COMPLETE")
    print(f"Output: {output_path}")
    print("=" * 65)




         
              
        
